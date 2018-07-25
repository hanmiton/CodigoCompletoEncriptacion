import sys
import math
import random
import openpyxl
import xlsxwriter
from time import time

LAMBDA = 16 #security parameter
N = LAMBDA	
P = LAMBDA ** 2
Q = LAMBDA ** 5



def principal(m1,m2):
	numsA = []
	numsB = []
	aux = []
	numsAEncrypt = []
	numsBEncrypt = []
	keys= []
	doc = openpyxl.load_workbook('cifrado.xlsx')
	doc.get_sheet_names()
	hoja = doc.get_sheet_by_name('Hoja1')
	prueba = []
	i=2
	for fila in hoja.rows:
		print "a"
		for columna in fila:
			#print int(columna.value)
			print columna.coordinate, columna.value
			boln1 = bin(int(columna.value))
			print boln1
			
			'''boln2 = bin(0)

			if(len(boln1) > len(boln2)):
				for i in range(0, len(boln1) - len(boln2)):
					aux.append(0)
				boln2 = aux + boln2
				del aux[:]
			else:
				for i in range(0, len(boln2) - len(boln1)):
					aux.append(0)
				boln1 = aux + boln1
				del aux[:]
			print boln1
			print boln2	
			key = map(keygen,boln1)
			print key 
			boln1Encrypt = map(encrypt,key,boln1)
			boln2Encrypt = map(encrypt,key,boln2)
			print boln1Encrypt
			print boln2Encrypt
			sumEncrypt = map(add,boln1Encrypt,boln2Encrypt)
			print sumEncrypt
			strEncriptSum = ''.join(str(e) for e in sumEncrypt)
			print strEncriptSum

			workbook = xlsxwriter.Workbook('enc1.xlsx')
			worksheet = workbook.add_worksheet()
			celda = 'A' + repr(i)
			celda2 = 'B' + repr(i)
			i = i + 1 
			worksheet.write(celda, strEncriptSum)
			worksheet.write(celda2, str(len(sumEncrypt) ))
			workbook.close()
			print strEncriptSum
			numsAEncrypt.append(boln1Encrypt)
			numsBEncrypt.append(boln2Encrypt)
			'''
	return 1		
	
	
	'''sumEncrypt = []
	mulEnctypt = []
	res = []
	
	
	
	mulEnctypt = map(mult,boln1Encrypt, boln2Encrypt)
	
	resSuma = map (decrypt, key, sumEncrypt)
	strSuma = ''.join(str(e) for e in resSuma)


	workbook = xlsxwriter.Workbook('enc1.xlsx')
	worksheet = workbook.add_worksheet()
	i=2
	celda = 'A' + repr(i)
	celda2 = 'B' + repr(i)
	worksheet.write(celda, strEncriptSum)
	worksheet.write(celda2, str(len(sumEncrypt) ))
	workbook.close()

	decSuma = int(strSuma, 2)

	#start_time = time()
	resMult = map (decrypt, key, mulEnctypt)
	#elapsed_time = time() - start_time

	#return elapsed_time 

	strMult = ''.join(str(e) for e in resMult)
	decMult = int(strMult, 2)
	
	return decSuma'''




		
def quot(z, p):
        # http://stackoverflow.com/questions/3950372/round-with-integer-division
        return (z + p // 2) // p
        
def mod(z, p):
        return z - quot(z,p) * p

def keygen(n):
	key = random.getrandbits(P)
	while(key % 2 == 0):
		key = random.getrandbits(P)
	return key

def encrypt(key, aBit):
	q = random.getrandbits(Q)
	m_a = 2 * random.getrandbits(N - 1)
	c = key * q + m_a + aBit
	return c

def decrypt(key, cipherText):
	return mod(cipherText, key) % 2

def add(cipherText1, cipherText2):
	return cipherText1 + cipherText2

def mult(cipherText1, cipherText2):
	return cipherText1 * cipherText2

def bin(numero):
	binario = ""
	listaN = []
	listaRn = []
	if (numero >0):
		while (numero >0):
			if(numero % 2 ==0):
				listaN.append(0)
				binario="0"+binario
			else:
				listaN.append(1)
				binario = "1"+ binario
			numero = int (math.floor(numero/2))
	else:
		if (numero ==0):
			listaN.append(0)
			return listaN
		else:
			return " no se pudo convertir el numero. ingrese solo numeros positivos"	
	for i in reversed(listaN):
		listaRn.append(i)
	return listaRn

if __name__ == '__main__':
	principal(m1,m2)